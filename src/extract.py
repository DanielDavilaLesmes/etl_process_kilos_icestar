# src/extract.py

import os
import re 
from openpyxl import load_workbook
from src.config import COLUMNAS_CANTIDAD_BRUTA, CABECERA_FECHA

# Patrón de expresión regular para encontrar cualquier VTA###
VTA_PATTERN = re.compile(r'(VTA\d{3})', re.IGNORECASE)


def get_client_name(worksheet):
    """
    Busca dinámicamente el nombre del cliente en la columna B 
    al lado de la etiqueta 'Cliente' en la columna A, sin importar la fila.
    """
    # Buscamos en las primeras 10 filas y solo en las columnas A y B (max_col=2)
    for row in worksheet.iter_rows(min_row=1, max_row=10, max_col=2, values_only=False):
        
        label_cell = row[0] # Columna A
        value_cell = row[1] # Columna B
        
        if label_cell.value is not None:
            # Limpiamos y estandarizamos la etiqueta
            label = str(label_cell.value).strip().upper().replace(":", "") 
            
            if label == "CLIENTE":
                if value_cell.value is not None:
                    # ¡Cliente encontrado! Devolvemos el valor de la columna B.
                    return str(value_cell.value).strip()
                else:
                    return "CLIENTE_NO_ESPECIFICADO"
                    
    return "CLIENTE_DESCONOCIDO"


def _extract_section_table(worksheet, start_row, client_name, file_name, section_title, sheet_name): # ⬅️ RECIBE sheet_name
    """
    Función auxiliar para extraer los datos de una sección VTA dinámica.
    """
    
    extracted_data = []
    header_row_index = None
    header_row_values = None
    
    # 1. Búsqueda de la fila de Cabecera ('Fecha')
    for i in range(1, 6): # Buscamos hasta 5 filas debajo del título VTA
        current_row_index = start_row + i 
        
        try:
            current_row_raw_values = [cell.value for cell in worksheet[current_row_index]]
            current_row_clean_values = [str(val).strip() if val is not None else None 
                                        for val in current_row_raw_values]

            # Buscamos la columna de fecha
            if CABECERA_FECHA in current_row_clean_values: 
                header_row_index = current_row_index
                header_row_values = current_row_raw_values 
                break
        except Exception:
            continue
    
    if header_row_index is None:
        return []

    # 2. Mapeo de Índices de Cabeceras
    header_map = {}
    
    # Mapeo de la fecha
    fecha_index = next((i for i, v in enumerate(header_row_values) 
                        if str(v).strip() == CABECERA_FECHA), None)
    if fecha_index is not None:
        header_map[CABECERA_FECHA] = fecha_index

    # Mapeo de TODAS las posibles columnas de cantidad (Cargue, Descargue, Cantidad, etc.)
    for header_bruta in COLUMNAS_CANTIDAD_BRUTA.keys():
        for col_index, header_value in enumerate(header_row_values):
            if str(header_value).strip() == header_bruta:
                header_map[header_bruta] = col_index
                
# 3. Extracción de Datos
    start_data_row = header_row_index + 1
    
    for row_index in range(start_data_row, worksheet.max_row + 1):
        try:
            row_values = [cell.value for cell in worksheet[row_index]]
            
            if row_values[0] is None and all(v is None for v in row_values[1:5]):
                break
            
            fecha_col_index = header_map.get(CABECERA_FECHA)
            fecha_bruta = row_values[fecha_col_index] if fecha_col_index is not None else None
            
            if fecha_bruta is None:
                continue

            for header_bruta, sub_tipo_movimiento in COLUMNAS_CANTIDAD_BRUTA.items():
                
                kilos_col_index = header_map.get(header_bruta)
                
                if kilos_col_index is not None:
                    kilos_brutos = row_values[kilos_col_index]

                    if kilos_brutos is not None and isinstance(kilos_brutos, (int, float)) and kilos_brutos > 0:
                        
                        row_data = {
                            'FECHA_MOVIMIENTO': fecha_bruta,
                            'CLIENTE': client_name,
                            'TIPO_MOVIMIENTO': section_title, 
                            'SUBTIPO_MOVIMIENTO': sub_tipo_movimiento, 
                            'CANTIDAD_MOVIMIENTO': kilos_brutos,
                            'ORIGEN_SECCION': header_bruta, 
                            'ORIGEN_HOJA': sheet_name, # ⬅️ AÑADIDA
                            'FUENTE_ARCHIVO': file_name
                        }
                        extracted_data.append(row_data)

        except Exception:
            continue

    return extracted_data


def extract_data_from_workbook(file_path):
    """
    Procesa un único archivo Excel RECORRIENDO TODAS LAS HOJAS 
    para extraer todas las secciones que contengan el patrón VTA###.
    """
    all_extracted_rows = []
    file_name = os.path.basename(file_path)

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        client_name = get_client_name(workbook.active)
        print(f"    -> Cliente identificado: {client_name}")
        
        for sheet_name in workbook.sheetnames: # ⬅️ USAMOS sheet_name
            worksheet = workbook[sheet_name]
            
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                section_title_raw = str(row[0]).strip() if row[0] is not None else None
                
                if section_title_raw:
                    match = VTA_PATTERN.search(section_title_raw)
                    
                    if match:
                        section_title = section_title_raw
                        
                        # PASAMOS sheet_name a la función auxiliar
                        section_data = _extract_section_table(
                            worksheet=worksheet,
                            start_row=row_index, 
                            client_name=client_name,
                            file_name=file_name,
                            section_title=section_title,
                            sheet_name=sheet_name # ⬅️ PASAMOS sheet_name
                        )
                        
                        all_extracted_rows.extend(section_data)

    except FileNotFoundError:
        print(f"Error: Archivo no encontrado en la ruta: {file_path}")
    except Exception as e:
        print(f"Error CRÍTICO al procesar el archivo {file_name}: {e}")
    finally:
        if 'workbook' in locals():
            workbook.close()
            
    return all_extracted_rows